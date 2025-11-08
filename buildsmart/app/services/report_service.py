"""
Report generation service for creating PDF and Excel reports.

This service provides methods for generating various types of reports
including sales reports, order reports, and analytics reports.
"""
from datetime import datetime, date
from decimal import Decimal
from typing import Dict, List, Optional
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from app.services.analytics_service import AnalyticsService


class ReportService:
    """Service for generating reports in various formats."""
    
    @staticmethod
    def generate_sales_report(start_date: Optional[date] = None, 
                              end_date: Optional[date] = None,
                              shop_id: Optional[int] = None,
                              format: str = 'pdf') -> BytesIO:
        """
        Generate a sales report.
        
        Args:
            start_date: Start date for the report
            end_date: End date for the report
            shop_id: Optional shop ID to filter
            format: Report format ('pdf' or 'excel')
        
        Returns:
            BytesIO buffer containing the report
        """
        if format == 'pdf':
            return ReportService._generate_sales_pdf(start_date, end_date, shop_id)
        elif format == 'excel':
            return ReportService._generate_sales_excel(start_date, end_date, shop_id)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    @staticmethod
    def _generate_sales_pdf(start_date: Optional[date], end_date: Optional[date],
                            shop_id: Optional[int]) -> BytesIO:
        """Generate PDF sales report."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Title
        title = Paragraph("Sales Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Date range
        if start_date and end_date:
            date_text = f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
        else:
            date_text = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        date_para = Paragraph(date_text, styles['Normal'])
        elements.append(date_para)
        elements.append(Spacer(1, 0.3*inch))
        
        # Sales overview
        overview = AnalyticsService.get_sales_overview(start_date, end_date)
        
        overview_data = [
            ['Metric', 'Value'],
            ['Total Sales', f"₦{overview['total_sales']:,.2f}"],
            ['Total Orders', f"{overview['total_orders']}"],
            ['Completed Orders', f"{overview['completed_orders']}"],
            ['Pending Orders', f"{overview['pending_orders']}"],
            ['Average Order Value', f"₦{overview['avg_order_value']:,.2f}"],
            ['Total Discounts', f"₦{overview['total_discounts']:,.2f}"],
            ['Total Taxes', f"₦{overview['total_taxes']:,.2f}"]
        ]
        
        overview_table = Table(overview_data, colWidths=[3*inch, 2*inch])
        overview_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(overview_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Top products
        top_products = AnalyticsService.get_top_products(limit=10, start_date=start_date, end_date=end_date)
        
        if top_products:
            products_data = [['Product', 'Shop', 'Quantity', 'Revenue']]
            for product in top_products:
                products_data.append([
                    product['product_name'],
                    product['shop_name'],
                    str(product['total_quantity']),
                    f"₦{product['total_revenue']:,.2f}"
                ])
            
            products_table = Table(products_data, colWidths=[2.5*inch, 1.5*inch, 1*inch, 1.5*inch])
            products_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (2, 1), (3, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(Paragraph("Top Products", styles['Heading2']))
            elements.append(Spacer(1, 0.1*inch))
            elements.append(products_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Top shops
        top_shops = AnalyticsService.get_top_shops(limit=10, start_date=start_date, end_date=end_date)
        
        if top_shops:
            shops_data = [['Shop', 'Revenue', 'Orders', 'Avg Order Value']]
            for shop in top_shops:
                shops_data.append([
                    shop['shop_name'],
                    f"₦{shop['total_revenue']:,.2f}",
                    str(shop['total_orders']),
                    f"₦{shop['avg_order_value']:,.2f}"
                ])
            
            shops_table = Table(shops_data, colWidths=[2.5*inch, 1.5*inch, 1*inch, 1.5*inch])
            shops_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(Paragraph("Top Shops", styles['Heading2']))
            elements.append(Spacer(1, 0.1*inch))
            elements.append(shops_table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def _generate_sales_excel(start_date: Optional[date], end_date: Optional[date],
                               shop_id: Optional[int]) -> BytesIO:
        """Generate Excel sales report."""
        wb = Workbook()
        
        # Overview sheet
        ws_overview = wb.active
        ws_overview.title = "Sales Overview"
        
        # Title
        ws_overview['A1'] = "Sales Report"
        ws_overview['A1'].font = Font(size=16, bold=True)
        ws_overview.merge_cells('A1:B1')
        
        # Date range
        if start_date and end_date:
            ws_overview['A2'] = f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
        else:
            ws_overview['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        
        # Overview data
        overview = AnalyticsService.get_sales_overview(start_date, end_date)
        
        ws_overview['A4'] = "Metric"
        ws_overview['B4'] = "Value"
        ws_overview['A4'].font = Font(bold=True)
        ws_overview['B4'].font = Font(bold=True)
        
        row = 5
        metrics = [
            ('Total Sales', f"₦{overview['total_sales']:,.2f}"),
            ('Total Orders', overview['total_orders']),
            ('Completed Orders', overview['completed_orders']),
            ('Pending Orders', overview['pending_orders']),
            ('Average Order Value', f"₦{overview['avg_order_value']:,.2f}"),
            ('Total Discounts', f"₦{overview['total_discounts']:,.2f}"),
            ('Total Taxes', f"₦{overview['total_taxes']:,.2f}")
        ]
        
        for metric, value in metrics:
            ws_overview[f'A{row}'] = metric
            ws_overview[f'B{row}'] = value
            row += 1
        
        # Top products sheet
        ws_products = wb.create_sheet("Top Products")
        top_products = AnalyticsService.get_top_products(limit=10, start_date=start_date, end_date=end_date)
        
        ws_products['A1'] = "Product"
        ws_products['B1'] = "Shop"
        ws_products['C1'] = "Quantity"
        ws_products['D1'] = "Revenue"
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col in ['A1', 'B1', 'C1', 'D1']:
            ws_products[col].fill = header_fill
            ws_products[col].font = header_font
        
        row = 2
        for product in top_products:
            ws_products[f'A{row}'] = product['product_name']
            ws_products[f'B{row}'] = product['shop_name']
            ws_products[f'C{row}'] = product['total_quantity']
            ws_products[f'D{row}'] = product['total_revenue']
            row += 1
        
        # Top shops sheet
        ws_shops = wb.create_sheet("Top Shops")
        top_shops = AnalyticsService.get_top_shops(limit=10, start_date=start_date, end_date=end_date)
        
        ws_shops['A1'] = "Shop"
        ws_shops['B1'] = "Revenue"
        ws_shops['C1'] = "Orders"
        ws_shops['D1'] = "Avg Order Value"
        
        for col in ['A1', 'B1', 'C1', 'D1']:
            ws_shops[col].fill = header_fill
            ws_shops[col].font = header_font
        
        row = 2
        for shop in top_shops:
            ws_shops[f'A{row}'] = shop['shop_name']
            ws_shops[f'B{row}'] = shop['total_revenue']
            ws_shops[f'C{row}'] = shop['total_orders']
            ws_shops[f'D{row}'] = shop['avg_order_value']
            row += 1
        
        # Adjust column widths
        for ws in [ws_overview, ws_products, ws_shops]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def generate_shop_report(shop_id: int, start_date: Optional[date] = None,
                            end_date: Optional[date] = None, format: str = 'pdf') -> BytesIO:
        """
        Generate a shop-specific report.
        
        Args:
            shop_id: Shop ID
            start_date: Start date for the report
            end_date: End date for the report
            format: Report format ('pdf' or 'excel')
        
        Returns:
            BytesIO buffer containing the report
        """
        if format == 'pdf':
            return ReportService._generate_shop_pdf(shop_id, start_date, end_date)
        elif format == 'excel':
            return ReportService._generate_shop_excel(shop_id, start_date, end_date)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    @staticmethod
    def _generate_shop_pdf(shop_id: int, start_date: Optional[date], end_date: Optional[date]) -> BytesIO:
        """Generate PDF shop report."""
        from app.models import Shop
        
        shop = Shop.query.get(shop_id)
        if not shop:
            raise ValueError(f"Shop {shop_id} not found")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Title
        title = Paragraph(f"Shop Report: {shop.name}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Date range
        if start_date and end_date:
            date_text = f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
        else:
            date_text = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        date_para = Paragraph(date_text, styles['Normal'])
        elements.append(date_para)
        elements.append(Spacer(1, 0.3*inch))
        
        # Shop analytics
        analytics = AnalyticsService.get_shop_analytics(shop_id, start_date, end_date)
        
        overview_data = [
            ['Metric', 'Value'],
            ['Total Orders', f"{analytics['total_orders']}"],
            ['Total Revenue', f"₦{analytics['total_revenue']:,.2f}"],
            ['Average Order Value', f"₦{analytics['avg_order_value']:,.2f}"]
        ]
        
        overview_table = Table(overview_data, colWidths=[3*inch, 2*inch])
        overview_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(overview_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Top products
        if analytics['top_products']:
            products_data = [['Product', 'Quantity', 'Revenue']]
            for product in analytics['top_products']:
                products_data.append([
                    product['product_name'],
                    str(product['quantity']),
                    f"₦{product['revenue']:,.2f}"
                ])
            
            products_table = Table(products_data, colWidths=[3.5*inch, 1.5*inch, 2*inch])
            products_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(Paragraph("Top Products", styles['Heading2']))
            elements.append(Spacer(1, 0.1*inch))
            elements.append(products_table)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def _generate_shop_excel(shop_id: int, start_date: Optional[date], end_date: Optional[date]) -> BytesIO:
        """Generate Excel shop report."""
        from app.models import Shop
        
        shop = Shop.query.get(shop_id)
        if not shop:
            raise ValueError(f"Shop {shop_id} not found")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Shop Report"
        
        # Title
        ws['A1'] = f"Shop Report: {shop.name}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')
        
        # Date range
        if start_date and end_date:
            ws['A2'] = f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
        else:
            ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        
        # Analytics
        analytics = AnalyticsService.get_shop_analytics(shop_id, start_date, end_date)
        
        ws['A4'] = "Metric"
        ws['B4'] = "Value"
        ws['A4'].font = Font(bold=True)
        ws['B4'].font = Font(bold=True)
        
        row = 5
        metrics = [
            ('Total Orders', analytics['total_orders']),
            ('Total Revenue', f"₦{analytics['total_revenue']:,.2f}"),
            ('Average Order Value', f"₦{analytics['avg_order_value']:,.2f}")
        ]
        
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            row += 1
        
        # Top products
        row += 2
        ws[f'A{row}'] = "Top Products"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        ws[f'A{row}'] = "Product"
        ws[f'B{row}'] = "Quantity"
        ws[f'C{row}'] = "Revenue"
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].font = header_font
        
        row += 1
        for product in analytics['top_products']:
            ws[f'A{row}'] = product['product_name']
            ws[f'B{row}'] = product['quantity']
            ws[f'C{row}'] = product['revenue']
            row += 1
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

